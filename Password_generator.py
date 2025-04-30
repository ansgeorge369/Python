# This Python program is a simple password validation and storage tool.
# It prompts the user to enter a password with certain criteria:
# - At least 6 characters long
# - Contains at least one uppercase letter
# - Contains at least one digit
# - Does NOT contain special characters from the defined set
# The password is stored in an Excel file ("Book1.xlsx") if it passes all conditions
# and is not already present. The user is given a limited number of chances to enter
# a valid password. Visual feedback is provided with colored terminal messages and emojis.
# Libraries used: openpyxl (for Excel handling), os (for file check), emoji (for visual feedback)

# Importing the emoji module to display emojis in output
import emoji

# Importing load_workbook and Workbook to read/write Excel files
from openpyxl import load_workbook, Workbook

# Importing os to check if a file exists
import os

# Function to check if the given password already exists in the Excel file
def password_exists(password, filename):
    # Check if the file exists
    if not os.path.exists(filename):
        return False
    # Load the Excel workbook
    workbook = load_workbook(filename)
    # Select the active worksheet
    sheet = workbook.active
    # Iterate through each row and check if the password is already saved
    for row in sheet.iter_rows(values_only=True):
        if password in row:
            return True
    # Return False if password not found
    return False

# Function to save a valid password to the Excel file
def save_password_to_excel(password, filename):
    # If the file does not exist, create a new workbook and sheet
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Passwords"
        # Add header to the new Excel sheet
        sheet.append(["Password"])
    else:
        # Load the existing workbook and active sheet
        workbook = load_workbook(filename)
        sheet = workbook.active
    # Append the new password to the sheet
    sheet.append([password])
    # Save the workbook
    workbook.save(filename)

# Main function to prompt user and validate password
def my_pass(chances):
    # Define the name of the Excel file to store passwords
    filename = "Book1.xlsx"
    # Loop while chances remain
    while chances > 0:
        # Ask the user to input a new password
        password = input("Enter the new password: ")
        # Get the length of the password
        length = len(password)
        # Check if the password contains at least one uppercase letter
        upper_case = any(char.isupper() for char in password)
        # Check if the password contains at least one digit
        digits = any(char.isdigit() for char in password)
        # Define a set of disallowed special characters
        special_char = "!@#$%^&*()_+"
        # Check if the password contains any disallowed special characters
        has_special = any(char in special_char for char in password)

        # Check for minimum length
        if length <= 5:
            print("\033[91mPlease enter at least 6 characters\033[0m")
        # Reject if special characters are used
        elif has_special:
            print("\033[91mPlease avoid special characters\033[0m")
        # Reject if no uppercase letter is present
        elif not upper_case:
            print("\033[91mPlease use at least 1 uppercase letter\033[0m")
        # Reject if no digit is present
        elif not digits:
            print("\033[91mPlease use at least 1 digit\033[0m")
        # Reject if the password already exists in the Excel file
        elif password_exists(password, filename):
            print("\033[93mPassword already exists. Please use a different one.\033[0m")
        # If all conditions are met, accept and save the password
        else:
            print(emoji.emojize("\033[92mPassword is set\033[0m :white_check_mark:"))
            save_password_to_excel(password, filename)
            return
        # Reduce the number of remaining chances
        chances -= 1
        # Notify the user of remaining chances
        print(f"Chances left: {chances}")

    # Notify the user after exhausting all chances
    print("Try after sometime.")

# Start the password creation with 4 chances
my_pass(4)
